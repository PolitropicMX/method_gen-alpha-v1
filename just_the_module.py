   # ZONA DE IMPORTACION DE CODIGO
from datetime import date
import json
from openpyxl import *
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
import openpyxl
import tkinter
import os
import sys
import glob
from natsort import natsorted

def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS2 
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)



# COMENTARIOS ACERCA DEL CODIGO:
# SI SE AÑADE MAS FILAS EN LA PRIMERA PARTE, RECORDAR CAMBIAR TODOS LOS 'x#' EJEMPLO (A1,B2) DE TODAS LAS LINEAS QUE REFERENCIEN

# Debemos numerar cada paso para los operadores



class Save_module:
    def __init__(self):
        pass
    def stylin(self,cell,f="Calibri",fc="000000",fs=11,fgc="ffffff",h="center",v="center",wt=True,bo=False):
        # PARAMETERS: CELL:'A2', FC = FONT COLOR, FS = FONT SIZE, FGC = FOREGROUND COLOR, , h = Horizontal, v = Vertical, wt = wrap_text
        self.ws[cell].font = Font(name=f,color=fc,size=fs,b=bo)
        self.ws[cell].fill = PatternFill("solid", fgColor=fgc)
        self.ws[cell].alignment = Alignment(horizontal=h, vertical=v,wrap_text= wt)

    def combine(self,cell1,cell2):
        self.ws.merge_cells(cell1+":"+cell2)

    # HOT NEW--------------------------------------------------------------------------------------------------------------------------------------
    def banner_antes_func(self,banner_antes,fila,n):# FUNCION (BOOLEANO,FILA,CONTADOR)
        if banner_antes and n == 1: #antes del titulo ( SI NO ES FALSO Y N = 1)
                            #   ANTES DEL ENCABEZADO
            parte_superior_2 = [# TEXTO
                    ['A'+str(fila),'No. Inven y Descripcion'],
                    ['B'+str(fila),'PICTOGRAMAS'],
                    ['E'+str(fila),'INDICACION'],
                    ['F'+str(fila),'ANTES DE USAR...']
                ]
            # ESTILOS
            self.combine("B"+str(fila),"D"+str(fila))
                # No. Inven y Descripcion
            self.stylin('A'+str(fila),fc="ffffff",fgc="000000" )
                # PICTOGRAMAS
            self.stylin('B'+str(fila),fc="ffffff",fgc="000000" )
                # Indicacion
            self.stylin('E'+str(fila),fc="ffffff",fgc="000000" )
                # Antes de usar
            self.stylin('F'+str(fila),fc="ffffff",fgc="000000" )
            # PEGADO AL DOCUMENTO
            for i,j in enumerate(parte_superior_2):
                    # SE AÑADE EL TEXTO 
                self.ws[j[0]].value = j[1]
                    # SE ENCUADRA EL TEXTO
                self.ws[j[0]].alignment = Alignment(wrap_text= True)
            fila += 1# OCUPA UNA FILA
            return fila# SI SE OCUPA SUMA UNA FILA Y LA DEVUELVE
        return fila # SI NO, DEVUELVE LA MISMA FILA
                
    def active_module(self,nombre,Noinv,dicc):
        
        #   SE CARGAN LOS DATOS:DICCIONARIO
        datos_cargados = dicc
        #   SE CARGA EL LIBRO DE EXCEL
        self.wb = Workbook()
        #   SE SELECCIONA UNA HOJA DEL LIBRO DE EXCEL
        self.ws = self.wb.active
        #   SE CAMBIAN EL TAMAÑO DE LAS COLUMNAS
        #
        #       DISEÑO DE LA PAGINA
        #
        #   por cada unidad en esta linea son 7 px
        self.ws.column_dimensions['A'].width = 21
        self.ws.column_dimensions['B'].width = 8
        self.ws.column_dimensions['C'].width = 8
        self.ws.column_dimensions['D'].width = 8
        self.ws.column_dimensions['E'].width = 22
        self.ws.column_dimensions['F'].width = 22
            # SE CAMBIAN EL TAMAÑO DE LAS FILAS
        self.ws.row_dimensions[3].height = 50
        self.ws.row_dimensions[4].height = 50
            #Obtener la fecha actual del sistema
        Fecha_elaboracion = str(date.today())
            # Recomencdaciones generales y 
        parte_superior = [
                   ['B1','Metodo de Fabricacion Estandar'],
                   ['F1','Documento Controlado'],
                   ['A2','PETRA'], # 0
                   ['B2', nombre+' '+Noinv], # 1
                   ['F2','Fecha de elaboracion '+Fecha_elaboracion], # 2
                   ['A4','Equipo a utilizar'],# 3
##                   ['B2','POLIPASTO, TINA CHICA DE ACERO INOX., DISPERSOR, BASCULAS, PLATAFORMA( GALLINA), PATIN'],# 4
                   ['E3','ANTES DE INICIAR, LIMPIE PERFECTAMENTE EL EQUIPO QUE VAYA A UTILIZAR, Y MANTENGA SU LUGAR DE TRABAJO LIMPIO, TENGA UNA FRANELA PARA EVITAR ENSUCIAR'],# 5
                   ['A3','Verifique que tenga todos los insumos indicados en la orden de producción, y que se encuentren debidamente identificados y APROBADOS'],# 6
                   ['E4','Si en su orden encuentra otro numero de inventario diferente al MFE, consulte con el SPR antes de iniciar su proceso	']
                   ]
        #   ESTA PARTE IRA DESPUES DE LOS PRIMEROS BANNERS
        
        
        
            # COMBINAR CELDAS PARTE SUPERIOR
##        self.ws.merge_cells("B1:E1")
        self.combine("B1","E1")
        self.combine("B2","E2")
        self.combine("E3","F3")
        self.combine("A3","D3")
        self.combine("E4","F4")
        self.combine("B4","D4")        
##        self.ws.merge_cells("B2:D2")  ESTA NO, POR QUE UNA VEZ SELECCIONADO EL PRODUCTO SE AÑADIRÁ EL EQUIPO A  UTILIZAR

        
            # PICTOGRAMAS
        letra_picto = ['a','c','i','m','t','r']
            # TODAS LAS IMAGENES
        images = {'a':'images/a.png',# ATENCION
                  'c':'images/c.png',# CORROSIVO
                  'i':'images/i.png',# INFLAMABLE
                  'm':'images/m.png',# DAÑO AMBIENTAL
                  't':'images/t.png',# TOXICO
                  'r':'images/r.png',# MORTAL
                  'b1':'images/advertisement.png', # BANNER DE PREPARACION
                  'b2':'images/epp.png', # BANNER DE EPP
                  'b3':'images/orden.png', # BANNER DE "PESE, CALCULE, ANOTE, AÑADA, TIER/GUARDE"
                  'b4':'images/klipton.png', # BANNER DE "MEZCLAR EN 4 PARTES"
                  'b5':'images/advert_explo_607.png', # BANNER DE advert. de explosion
                  'b6':'images/bullton5085advert.png',# ADVERTENCIA DEL TDI
                  'b7':'images/consideraciones_ban.png',# CONSIDERACIONES BÁSICAS ANTES DE EMPEZAR EL PROCESO
                  'b8':'images/ADVERLIMPIEZA.png', # MANTENGA LIMPIO SU REACTOR
                  'b9':'images/pasarmuestra.png',# PASE MUESTRA A CALIDAD
                  'b10':'images/solicitar_fineza.png',# Solicitar Fineza/Finura
                  'b11':'images/pasarmuestra2.png',# PASE MUESTRA A CALIDAD bote sucio
                  'b12':'images/evitedispersor.png'# EVITE DEJAR FUNCIONANDO MUCHO TIEMPO EL DISPERSOR
                  }#C:\Users\Fernando.Lopez\Downloads\los_codigos\los_codigos\openpyxl\codigo_que_escribe_codigo\Method_gen\pruebasa.png  C:\Users\Fernando.Lopez\Downloads\los_codigos\los_codigos\openpyxl\codigo_que_escribe_codigo\Method_gen\pruebas
                    #C:/Users/Fernando.Lopez/Downloads/los_codigos/los_codigos/openpyxl/codigo_que_escribe_codigo/Method_gen/pruebas/
        # Images = {'a':'Ruta de a',
        #           'c':'Ruta de c',...}
            # CICLO FOR: AQUI SE CARGA A "PARTE SUPERIOR" AL ARCHIVO
        for i,j in enumerate(parte_superior):
                # SE AÑADE EL TEXTO 
            self.ws[j[0]].value = j[1]
                # SE ENCUADRA EL TEXTO
            self.ws[j[0]].alignment = Alignment(wrap_text= True)

        # SE ESTILIZA LA PARTE DE ARRIBA
            # // COLORES VERDE CLARO 08bd83
            # // COLOR VERDE MAS CLARO 2cf2b3
            # // COLORES VERDE PETRA 046546

        
            # SE ESTILIZA EL TITULO
                # Metodo de Fabricacion Estandar
        self.stylin('A1',fgc="000000")
        self.stylin('B1',f='arial',fc="ffffff",fs=14,fgc="000000")
                # Documento Confidencial
        self.stylin('F1',f='arial',fc="ffffff",fs=10,fgc="000000")
                # PETRA
        self.stylin('A2',f="Berlin Sans FB Demi",fc="046546",fs=20,fgc="ffffff",)
                # Titulo del Documento 7 # Color PETRA 046546
        self.stylin('B2',f="Antique Olive Roman",fc="ffffff", fs=20,fgc="046546")
                # Fecha de Elaboracion
        self.stylin('F2',f="Albertus MT")
            # Equipo a utilizar
        self.stylin('A3',fc="ffffff",fgc="08bd83")
            # ANTES DE INICIAR LIMPIE...
        self.stylin('E3',fc="ffffff", fs=9,fgc="08bd83" )
            # Verifique que tenga
        self.stylin('A4',fc="000000", fs=9,fgc="2cf2b3" )
        
        
        # AQUI YA SE CARGO LA INFORMACION DE LA PARTE SUPERIOR
        ############################################################################################################################################
        
        # A PARTIR DE AQUI SE CARGA LA INFORMACIÓN DE LO REACTIVOS
        fila = 5
##        for i,producto in enumerate(datos_cargados): 
##            print(f'{i}: {producto}')
        print('elija segun su NOMBRE:')
        op2 = nombre
            # SE SELECCIONA EL PAQUETE DE  INFORMACIÓN
        if op2 in datos_cargados:# SE COMPRUEBA SI EL STRING INGRESADO ESTA EN EL DICCIONARIO
            self.reactivos_hoja = []
                # FILA DONDE SE EMPIEZA A AGREGAR LA INFORMACION
            
            print('Si esta')
                # SE ACOMODAN LOS DATOS
                # EQUIPO A UTILIZAR
            equipo_a_usar = datos_cargados[op2]["Equipo a utilizar"]["Equipo a utilizar"]
            self.ws['B4'].value = equipo_a_usar
            self.ws.merge_cells("B3:D3")
            self.ws['B4'].alignment = Alignment(horizontal="center", vertical="center",wrap_text= True)
            self.envase_producto = datos_cargados[op2]["Envase"]["Envase_producto"]
            self.envase_control = datos_cargados[op2]["Envase"]["Envase_control"]
                # Se itera a traves de el producto seleccionado
                # reactivo es la llave del nivel 2
                # {key1: value1}
                # {Key1: {Key2:Value2}}
                # {Key1: {Key2: {Key3: value}}}
            n_paso = 1
            n_banner = 1
            n_reactivo = 1
            n_no_banner = 1
            banner_antes = True
            for i,reactivo in enumerate(datos_cargados[str(op2)].keys()):
                print(reactivo)
                
                # AQUI ES CUANDO SE CARGA LA INFORMACION DE LOS REACTIVOS
                    # NO SE ELIGE EL ULTIMO NI EL PENULTIMO ELEMENTO DE CADA PRODUCTO
                if i != len(datos_cargados[str(op2)].keys())-1 and i != len(datos_cargados[str(op2)].keys())-2:
                        # SE MUESTRA EL NUMERO DE ITERACION

                    # Es reactivo o paso?
                    if reactivo[:4] == "Paso":# LOS PRIMEROS 4 CARACTERES
                        
                        print(banner_antes)
                        fila = self.banner_antes_func(banner_antes,fila,n_no_banner)
                        banner_antes = False
                        n_no_banner += 1
                        print(f'{i}# Iteracion Paso')
                        paso = datos_cargados[op2][f"Paso {n_paso}"]["texto"]# Texto
                        self.reactivos_hoja.append(['A'+str(fila),paso])# tupla ("celda","texto")
                        self.ws.merge_cells('A'+str(fila)+':F'+str(fila))# combinar celdas
                        
                        n_paso += 1
                        fila += 1
                        
                    elif reactivo[:6] == "Banner":
                        print(f'{i}# Iteracion y fila {fila}')
                        ruta = images[datos_cargados[op2][f'Banner {n_banner}']]
                        self.ws.add_image(Image(ruta),anchor='A'+str(fila))
                        n_banner += 1
                        fila += 3
                    elif reactivo[:6] == "Inform":
                        pass
                    elif reactivo[0] == "1" or reactivo[0] == "2":
                        print(banner_antes)
                        fila = self.banner_antes_func(banner_antes,fila,n_no_banner)
                        banner_antes = False
                        n_no_banner += 1
                        n_reactivo += 1
                        print(f'{i}# Iteracion ')

                            # SE AÑADE LA INFORMACION DE LA BASE DE DATOS AL ARRAY
                                # No inventario
                        self.reactivos_hoja.append(['A'+str(fila),reactivo])
                                # DESCRIPCION
                        print(f'op2 es {op2} y reactivo es {reactivo}')
                        descripcion = datos_cargados[op2][reactivo]['descripcion']
                        self.reactivos_hoja.append(['A'+str(fila+1),descripcion])
                                # INDICACION
                        indicacion = datos_cargados[op2][reactivo]['indicacion']
                        self.reactivos_hoja.append(['E'+str(fila),indicacion])
                                # REVISION DE USO
                        revision = datos_cargados[op2][reactivo]['revision']
                        self.reactivos_hoja.append(['F'+str(fila),revision])
                            # SE MODIFICAN LAS CELDAS (ESTILOS)
                        self.ws.merge_cells('A'+str(fila+1)+':A'+str(fila+2))# se combinan (DESCRIPCION)
                        self.ws.merge_cells('E'+str(fila)+':E'+str(fila+2))# se combinan (INDICACION)
                        self.ws.merge_cells('F'+str(fila)+':F'+str(fila+2))# se combinan (REVISION)
                        self.ws.row_dimensions[fila+1].height = 30

                            # SE AÑADEN LOS PICTOGRAMAS
                        peligro = datos_cargados[op2][reactivo]['peligro']
                        if peligro != '0':
                            for k,pictograma in enumerate(peligro):
                                ruta = images[pictograma]
##                                    print(ruta)
                                self.ws.add_image(Image(ruta),anchor=chr(66+k)+str(fila))
                                
                                if pictograma == 'm':
                                    self.ws[chr(66+k)+str(fila+2)].value = 'Daño Ambiental'
                                elif pictograma == 'a':
                                    self.ws[chr(66+k)+str(fila+2)].value = 'Atencion'
                                elif pictograma == 'c':
                                    self.ws[chr(66+k)+str(fila+2)].value = 'Corrosivo'
                                elif pictograma == 't':
                                    self.ws[chr(66+k)+str(fila+2)].value = 'Toxico'
                                elif pictograma == 'i':
                                    self.ws[chr(66+k)+str(fila+2)].value = 'Inflamable'
                                    
                                self.stylin(chr(66+k)+str(fila+2),fs=8)
                                self.stylin(chr(66)+str(fila),fgc="046546")
                                self.stylin(chr(66)+str(fila+1),fgc="046546")
                                self.stylin(chr(67)+str(fila),fgc="046546")
                                self.stylin(chr(67)+str(fila+1),fgc="046546")
                                self.stylin(chr(68)+str(fila),fgc="046546")
                                self.stylin(chr(68)+str(fila+1),fgc="046546")
                                
                                
                        fila += 3
        
                
                # SE AÑADEN AL EXCEL LOS REACTIVOS
            solo_reac = 1
            for i,anadir in enumerate(self.reactivos_hoja):
                self.ws[anadir[0]].value = anadir[1]
                self.ws[anadir[0]].alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)# SE ENCUADRA EL TEXTO
                    # EL NOMBRE DEL REACTIVO SE RESALTA

                if anadir[0][0] == 'A' and solo_reac == 1 and (anadir[1][0] == '1' or anadir[1][0] == '2'):
##                        print(f' funciono ? {anadir[0]}')
                    self.ws[anadir[0]].font = Font(name="Albertus MT",color="ffffff", size=12)
                    self.ws[anadir[0]].fill = PatternFill("solid", fgColor="046546")
                    solo_reac = 3
##                    elif anadir[1][0] != '1' or :
##                        self.ws.row_dimensions[int(anadir[0][1])].height = 40
##                        solo_reac = 1
                else:
                    solo_reac = 1
                if anadir[0][0] == 'A' and  not anadir[1][0] == '1':
                    
                    if len(anadir[1]) >= 85:
                        print(anadir[1])
                        self.ws.row_dimensions[int(anadir[0][1:])].height = 40
                    if len(anadir[1]) >= 160:
                        print(anadir[1])
                        self.ws.row_dimensions[int(anadir[0][1:])].height = 50
                        self.ws[anadir[0]].font= Font(size=10)
##                    print(f'anadir es {anadir[0]} y su valor {anadir[1]}')
                
        ####################################################################################################################################################3

                # INFORMACION A AÑADIR: PARTE INFERIOR
            parte_inferior = [['A'+str(fila),'TOMA DE MUESTRA'],
                              ['A'+str(fila+1),f'1)Tome una muestra en un envase de {self.envase_control} (Debe estar limpio y seco)'],
                              ['A'+str(fila+2),'2) Con un marcador, coloque el nombre del producto.'],
                              ['A'+str(fila+3),'3) Entregue la muestra al ICC junto con la orden de producción correspondiente.'],
                              ['D'+str(fila),'ENVASADO'],
                              ['D'+str(fila+1),'1) Verifique que la malla este limpia y no tenga residuos de otro material.'],
                              ['D'+str(fila+2),'2) Verifique que los envases no esten sucios y no tengan polvo.'],
                              ['D'+str(fila+3),'3) Al cerrar los envases verifique que esten bien cerrados.'],
                              ['F'+str(fila+1),f'Una vez aprobado su producto proceda a envasar en {self.envase_producto}. Envase filtrando su material con una malla de 150 micras y envase con agitación'],
                              ['A'+str(fila+5),'ELABORÓ'],
                              ['A'+str(fila+6),'REVISÓ'],
                              ['A'+str(fila+7),'REVISÓ'], # HOT NEW
                              ['A'+str(fila+8),'AUTORIZÓ'], # HOT NEW
                              ['B'+str(fila+5),'Fernando López V.'],
                              ['B'+str(fila+6),'Leonardo Sánchez M.'],
                              ['B'+str(fila+7),'Jocelyn López B.'], # HOT NEW
                              ['B'+str(fila+8),'José de Jesus Cruz M.'], # HOT NEW
                              ['E'+str(fila+5),'Diseño y Desarrollo'],
                              ['E'+str(fila+6),'Jefe de Producción'],
                              ['E'+str(fila+7),'Coordinadora de Seguridad e Higiene'], # HOT NEW
                              ['E'+str(fila+8),'Jefe de Laboratorio'], # HOT NEW
                              ['F'+str(fila+4),'Firma']
                              
                              ]
                # SE ESTILIZA LA HOJA: TAMAÑO DE LAS FILAS
            self.ws.row_dimensions[fila+1].height = 50
            self.ws.row_dimensions[fila+2].height = 30
            self.ws.row_dimensions[fila+3].height = 30
            self.ws.row_dimensions[fila+5].height = 25
            self.ws.row_dimensions[fila+6].height = 25
            self.ws.row_dimensions[fila+7].height = 25
            self.ws.row_dimensions[fila+8].height = 25
                # SE PEGA LA INFORMACION
            for i,j in enumerate(parte_inferior):
                self.ws[j[0]].value = j[1] # SE AÑADE EL TEXTO 
                self.ws[j[0]].alignment = Alignment(wrap_text= True)# SE ENCUADRA EL TEXTO
                if j[0][0] == 'A' and int(j[0][1]) < 4:
                    print(f"el valor : {j[0][1]}, y fila: {fila}")
                    self.ws[j[0]].fill = PatternFill("solid", fgColor="08bd83")
                elif j[0][0] == 'D':
                    self.ws[j[0]].fill = PatternFill("solid", fgColor="daeef3")
                # SE COMBINAN LAS CELDAS
            self.ws.merge_cells("A"+str(fila)+":C"+str(fila))
            self.ws.merge_cells("A"+str(fila+1)+":C"+str(fila+1))
            self.ws.merge_cells("A"+str(fila+2)+":C"+str(fila+2))
            self.ws.merge_cells("A"+str(fila+3)+":C"+str(fila+3))
            self.ws.merge_cells("D"+str(fila)+":F"+str(fila))
            self.ws.merge_cells("D"+str(fila+1)+":E"+str(fila+1))
            self.ws.merge_cells("D"+str(fila+2)+":E"+str(fila+2))
            self.ws.merge_cells("D"+str(fila+3)+":E"+str(fila+3))
            self.ws.merge_cells("F"+str(fila+1)+":F"+str(fila+3))
            self.ws.merge_cells("B"+str(fila+5)+":D"+str(fila+5))
            self.ws.merge_cells("B"+str(fila+6)+":D"+str(fila+6))
            self.ws.merge_cells("B"+str(fila+7)+":D"+str(fila+7))
            self.ws.merge_cells("B"+str(fila+8)+":D"+str(fila+8))
                # ESTILO DE CELDA
            self.stylin('A'+str(fila),f="Albertus MT",fc="ffffff",fs=14,fgc="033827")
            self.stylin('D'+str(fila),f="Albertus MT",fc="ffffff",fs=14,fgc="033827")
            self.stylin('F'+str(fila+1),fs=8)
            self.ws['F'+str(fila+5)].border = Border(bottom=Side(style='medium'))
            self.ws['F'+str(fila+6)].border = Border(bottom=Side(style='medium'))
            self.ws['F'+str(fila+7)].border = Border(bottom=Side(style='medium'))
            self.ws['F'+str(fila+8)].border = Border(bottom=Side(style='medium'))
            self.stylin('F'+str(fila+4),fc="000000",fs=14,bo = True)
            
            
            self.reactivos_hoja =  []
        else:# EL USUARIO SE EQUIVOCO Y NO ESTA EN EL DICCIONARIO
            print('no está')
            if dicc == "0":
                print("No funciono")
            else:
                datos_cargados = dicc
                self.reactivos_hoja = []
                fila = 6
                    # SE ACOMODAN LOS DATOS
                    # EQUIPO A UTILIZAR
                equipo_a_usar = datos_cargados[op2]["Equipo a utilizar"]["Equipo a utilizar"]
                self.ws['B4'].value = equipo_a_usar
                self.ws.merge_cells("B3:D3")
                self.ws['B4'].alignment = Alignment(horizontal="center", vertical="center",wrap_text= True)
                self.envase_producto = datos_cargados[op2]["Envase"]["Envase_producto"]
                self.envase_control = datos_cargados[op2]["Envase"]["Envase_control"]
                    # Se itera a traves de el producto seleccionado
                    # reactivo es la llave del nivel 2
                    # {key1: value1}
                    # {Key1: {Key2:Value2}}
                    # {Key1: {Key2: {Key3: value}}}
                n_paso = 1
                n_banner = 1
                for i,reactivo in enumerate(datos_cargados[str(op2)].keys()):
                    print(reactivo)
                    
                    # AQUI ES CUANDO SE CARGA LA INFORMACION DE LOS REACTIVOS
                        # NO SE ELIGE EL ULTIMO NI EL PENULTIMO ELEMENTO DE CADA PRODUCTO
                    if i != len(datos_cargados[str(op2)].keys())-1 and i != len(datos_cargados[str(op2)].keys())-2:
                            # SE MUESTRA EL NUMERO DE ITERACION

                        # Es reactivo o paso?
                        if reactivo[:4] == "Paso":# LOS PRIMEROS 4 CARACTERES
                            print(f'{i}# Iteracion Paso')
                            paso = datos_cargados[op2][f"Paso {n_paso}"]["texto"]
                            self.reactivos_hoja.append(['A'+str(fila),paso])
                            self.ws.merge_cells('A'+str(fila)+':F'+str(fila))
                            n_paso += 1
                            fila += 1
                        elif reactivo[:6] == "Banner":
                            ruta = images[datos_cargados[op2][f'Banner {n_banner}']]
                            self.ws.add_image(Image(ruta),anchor='A'+str(fila))
                            n_banner += 1
                            fila += 3
                        else:
                            print(f'{i}# Iteracion ')

                                # SE AÑADE LA INFORMACION DE LA BASE DE DATOS AL ARRAY
                                    # No inventario
                            self.reactivos_hoja.append(['A'+str(fila),reactivo])
                                    # DESCRIPCION
                            descripcion = datos_cargados[op2][reactivo]['descripcion']
                            self.reactivos_hoja.append(['A'+str(fila+1),descripcion])
                                    # INDICACION
                            indicacion = datos_cargados[op2][reactivo]['indicacion']
                            self.reactivos_hoja.append(['E'+str(fila),indicacion])
                                    # REVISION DE USO
                            revision = datos_cargados[op2][reactivo]['revision']
                            self.reactivos_hoja.append(['F'+str(fila),revision])
                                # SE MODIFICAN LAS CELDAS (ESTILOS)
                            self.ws.merge_cells('A'+str(fila+1)+':A'+str(fila+2))# se combinan (DESCRIPCION)
                            self.ws.merge_cells('E'+str(fila)+':E'+str(fila+2))# se combinan (INDICACION)
                            self.ws.merge_cells('F'+str(fila)+':F'+str(fila+2))# se combinan (REVISION)
                            self.ws.row_dimensions[fila+1].height = 30

                                # SE AÑADEN LOS PICTOGRAMAS
                            peligro = datos_cargados[op2][reactivo]['peligro']
                            if peligro != '0':
                                for k,pictograma in enumerate(peligro):
                                    ruta = images[pictograma]
    ##                                    print(ruta)
                                    self.ws.add_image(Image(ruta),anchor=chr(66+k)+str(fila))
                                    
                                    if pictograma == 'm':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Daño Ambiental'
                                    elif pictograma == 'a':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Atencion'
                                    elif pictograma == 'c':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Corrosivo'
                                    elif pictograma == 't':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Toxico'
                                    elif pictograma == 'i':
                                        self.ws[chr(66+k)+str(fila+2)].value = 'Inflamable'
                                        
                                    self.stylin(chr(66+k)+str(fila+2),fs=8)
                                    self.stylin(chr(66)+str(fila),fgc="046546")
                                    self.stylin(chr(66)+str(fila+1),fgc="046546")
                                    self.stylin(chr(67)+str(fila),fgc="046546")
                                    self.stylin(chr(67)+str(fila+1),fgc="046546")
                                    self.stylin(chr(68)+str(fila),fgc="046546")
                                    self.stylin(chr(68)+str(fila+1),fgc="046546")
                                    
                                    
                            fila += 3
            
                    
                    # SE AÑADEN AL EXCEL LOS REACTIVOS
                solo_reac = 1
                for i,anadir in enumerate(self.reactivos_hoja):
                    self.ws[anadir[0]].value = anadir[1]
                    self.ws[anadir[0]].alignment = Alignment(horizontal="center", vertical="center", wrap_text= True)# SE ENCUADRA EL TEXTO
                        # EL NOMBRE DEL REACTIVO SE RESALTA

                    if anadir[0][0] == 'A' and solo_reac == 1 and (anadir[1][0] == '1' or anadir[1][0] == '2'):
    ##                        print(f' funciono ? {anadir[0]}')
                        self.ws[anadir[0]].font = Font(name="Albertus MT",color="ffffff", size=12)
                        self.ws[anadir[0]].fill = PatternFill("solid", fgColor="046546")
                        solo_reac = 3
    ##                    elif anadir[1][0] != '1' or :
    ##                        self.ws.row_dimensions[int(anadir[0][1])].height = 40
    ##                        solo_reac = 1
                    else:
                        solo_reac = 1
                    if anadir[0][0] == 'A' and  not anadir[1][0] == '1':
                        
                        if len(anadir[1]) >= 85:
                            print(anadir[1])
                            self.ws.row_dimensions[int(anadir[0][1:])].height = 40
                        if len(anadir[1]) >= 160:
                            print(anadir[1])
                            self.ws.row_dimensions[int(anadir[0][1:])].height = 50
                            self.ws[anadir[0]].font= Font(size=10)
    ##                    print(f'anadir es {anadir[0]} y su valor {anadir[1]}')
                    
            ####################################################################################################################################################3

                    # INFORMACION A AÑADIR: PARTE INFERIOR
                parte_inferior = [['A'+str(fila),'TOMA DE MUESTRA'],
                                  ['A'+str(fila+1),f'1)Tome una muestra en un envase de {self.envase_control} (Debe estar limpio y seco)'],
                                  ['A'+str(fila+2),'2) Con un marcador, coloque el nombre del producto.'],
                                  ['A'+str(fila+3),'3) Entregue la muestra al ICC junto con la orden de producción correspondiente.'],
                                  ['D'+str(fila),'ENVASADO'],
                                  ['D'+str(fila+1),'1) Verifique que la malla este limpia y no tenga residuos de otro material.'],
                                  ['D'+str(fila+2),'2) Verifique que los envases no esten sucios y no tengan polvo.'],
                                  ['D'+str(fila+3),'3) Al cerrar los envases verifique que esten bien cerrados.'],
                                  ['F'+str(fila+1),f'Una vez aprobado su producto proceda a envasar en {self.envase_producto}. Envase filtrando su material con una malla de 150 micras y envase con agitación'],
                                  ['A'+str(fila+5),'ELABORÓ'],
                                  ['A'+str(fila+6),'REVISÓ'],
                                  ['A'+str(fila+7),'REVISÓ'], # HOT NEW
                                  ['A'+str(fila+8),'AUTORIZÓ'], # HOT NEW
                                  ['B'+str(fila+5),'Fernando López V.'],
                                  ['B'+str(fila+6),'Leonardo Sánchez M.'],
                                  ['B'+str(fila+7),'Jocelyn López B.'], # HOT NEW
                                  ['B'+str(fila+8),'José de Jesus Cruz M.'], # HOT NEW
                                  ['E'+str(fila+5),'Diseño y Desarrollo'],
                                  ['E'+str(fila+6),'Jefe de Producción'],
                                  ['E'+str(fila+7),'Coordinadora de Seguridad e Higiene'], # HOT NEW
                                  ['E'+str(fila+8),'Jefe de Laboratorio'], # HOT NEW
                                  ['F'+str(fila+4),'Firma']
                                  
                                  ]
                    # SE ESTILIZA LA HOJA: TAMAÑO DE LAS FILAS
                self.ws.row_dimensions[fila+1].height = 50
                self.ws.row_dimensions[fila+2].height = 30
                self.ws.row_dimensions[fila+3].height = 30
                self.ws.row_dimensions[fila+5].height = 25
                self.ws.row_dimensions[fila+6].height = 25
                self.ws.row_dimensions[fila+7].height = 25
                self.ws.row_dimensions[fila+8].height = 25
                    # SE PEGA LA INFORMACION
                for i,j in enumerate(parte_inferior):
                    self.ws[j[0]].value = j[1] # SE AÑADE EL TEXTO 
                    self.ws[j[0]].alignment = Alignment(wrap_text= True)# SE ENCUADRA EL TEXTO
                    if j[0][0] == 'A' and int(j[0][1]) < 4:
                        print(f"el valor : {j[0][1]}, y fila: {fila}")
                        self.ws[j[0]].fill = PatternFill("solid", fgColor="08bd83")
                    elif j[0][0] == 'D':
                        self.ws[j[0]].fill = PatternFill("solid", fgColor="daeef3")
                    # SE COMBINAN LAS CELDAS
                self.ws.merge_cells("A"+str(fila)+":C"+str(fila))
                self.ws.merge_cells("A"+str(fila+1)+":C"+str(fila+1))
                self.ws.merge_cells("A"+str(fila+2)+":C"+str(fila+2))
                self.ws.merge_cells("A"+str(fila+3)+":C"+str(fila+3))
                self.ws.merge_cells("D"+str(fila)+":F"+str(fila))
                self.ws.merge_cells("D"+str(fila+1)+":E"+str(fila+1))
                self.ws.merge_cells("D"+str(fila+2)+":E"+str(fila+2))
                self.ws.merge_cells("D"+str(fila+3)+":E"+str(fila+3))
                self.ws.merge_cells("F"+str(fila+1)+":F"+str(fila+3))
                self.ws.merge_cells("B"+str(fila+5)+":D"+str(fila+5))
                self.ws.merge_cells("B"+str(fila+6)+":D"+str(fila+6))
                self.ws.merge_cells("B"+str(fila+7)+":D"+str(fila+7))
                self.ws.merge_cells("B"+str(fila+8)+":D"+str(fila+8))
                    # ESTILO DE CELDA
                self.stylin('A'+str(fila),f="Albertus MT",fc="ffffff",fs=14,fgc="033827")
                self.stylin('D'+str(fila),f="Albertus MT",fc="ffffff",fs=14,fgc="033827")
                self.stylin('F'+str(fila+1),fs=8)
                self.ws['F'+str(fila+5)].border = Border(bottom=Side(style='medium'))
                self.ws['F'+str(fila+6)].border = Border(bottom=Side(style='medium'))
                self.ws['F'+str(fila+7)].border = Border(bottom=Side(style='medium'))
                self.ws['F'+str(fila+8)].border = Border(bottom=Side(style='medium'))
                self.stylin('F'+str(fila+4),fc="000000",fs=14,bo = True)
                
                
                self.reactivos_hoja =  []
        # AQUI SE CARGA LA ULTIMA PARTE DEL PROGRAMA
        

##      ULTIMA INSTRUCCION ANTES DE CERRAR EL ARCHIVO
        self.wb.save(f'{nombre}.xlsx')

        

### INICIALIZAMOS EL PROGRAMA ###
##clase_save = Save_module()
##clase_save.active_module("prueba2","XX-XXX")

